from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.db import IntegrityError
from django.urls import reverse
from django.contrib.auth import authenticate, login, logout
from .models import RegionMedCenter, User, Choices, Questions, Answer, Form, Responses, DateOfBirth, UserGender, Image, UserDesc, UserCity, UserMed
import json
import random
import string
import csv
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib import messages
import re
from django.http import HttpResponseForbidden
from django.shortcuts import render, get_object_or_404
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

@login_required
def add_medical_center(request):
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
        
    if request.method == "POST":
        region = request.POST.get('region')
        med_center = request.POST.get('med_center')
        address = request.POST.get('address')
        
        if region and med_center and address:
            RegionMedCenter.objects.create(
                region=region,
                med_center=med_center,
                address=address
            )
            messages.success(request, "Медицинский центр успешно добавлен")
            return HttpResponseRedirect(reverse("manage_medical_centers"))
            
    return render(request, "index/add_edit_medical_center.html", {
        "city_choices": UserCity.CITY_CHOICES
    })

@login_required
def edit_medical_center(request, center_id):
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
        
    center = get_object_or_404(RegionMedCenter, id=center_id)
    
    if request.method == "POST":
        region = request.POST.get('region')
        med_center = request.POST.get('med_center')
        address = request.POST.get('address')
        
        if region and med_center:
            center.region = region
            center.med_center = med_center
            center.address = address
            center.save()
            messages.success(request, "Медицинский центр успешно обновлен")
            return HttpResponseRedirect(reverse("manage_medical_centers"))
            
    return render(request, "index/add_edit_medical_center.html", {
        "center": center,
        "city_choices": UserCity.CITY_CHOICES
    })

@login_required
def manage_medical_centers(request):
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
        
    if request.method == "POST":
        action = request.POST.get('action')
        
        if action == 'add':
            region = request.POST.get('region')
            med_center = request.POST.get('med_center')
            address = request.POST.get('address')
            
            if region and med_center:
                RegionMedCenter.objects.create(
                    region=region,
                    med_center=med_center,
                    address=address or ''
                )
                messages.success(request, "Медицинский центр успешно добавлен")
                
        elif action == 'edit':
            center_id = request.POST.get('center_id')
            region = request.POST.get('region')
            med_center = request.POST.get('med_center')
            address = request.POST.get('address')
            
            if center_id:
                center = get_object_or_404(RegionMedCenter, id=center_id)
                center.region = region
                center.med_center = med_center
                center.address = address
                center.save()
                messages.success(request, "Медицинский центр успешно обновлен")
                
        elif action == 'delete':
            center_id = request.POST.get('center_id')
            if center_id:
                RegionMedCenter.objects.filter(id=center_id).delete()
                messages.success(request, "Медицинский центр успешно удален")
                
    med_centers = RegionMedCenter.objects.all().order_by('region', 'med_center')
    city_choices = UserCity.CITY_CHOICES
    
    return render(request, "index/manage_medical_centers.html", {
        "med_centers": med_centers,
        "city_choices": city_choices
    })

def update_med_center(request, user_id):
    user = get_object_or_404(User, id=user_id)
    med_centers = RegionMedCenter.objects.all().order_by('region', 'med_center')

    if request.method == 'POST':
        med_center_name = request.POST.get('med_center')
        if med_center_name:
            user_med_instance, created = UserMed.objects.get_or_create(
                user=user, defaults={'med_center': med_center_name}
            )

            if not created:
                user_med_instance.med_center = med_center_name
                user_med_instance.save()

            user.med_center = user_med_instance 
            user.save()
            messages.success(request, f"Медицинский центр для {user.username} был обновлен!")
            return redirect('update_med_center', user_id=user.id)
        else:
            messages.error(request, "Пожалуйста, выберите медицинский центр.")
    
    return render(request, 'index/user_detail.html', {
        'user': user,
        'med_centers': med_centers,
    })

@csrf_exempt
def delete_selected_responses(request):
    if request.method == "POST":
        data = json.loads(request.body)
        response_ids = data.get("response_ids", [])

        if response_ids:
            Responses.objects.filter(id__in=response_ids).delete()
            return JsonResponse({"success": True})

    return JsonResponse({"success": False}, status=400)

@login_required
def update_user_status(request, user_id):
    if not request.user.is_superuser:
        return HttpResponseForbidden()

    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        status = request.POST.get('status')
        
        if status == 'superuser':
            user.is_superuser = True
            user.is_staff = True
        elif status == 'staff':
            user.is_superuser = False
            user.is_staff = True
        else:
            user.is_superuser = False
            user.is_staff = False
        
        user.save()
        return redirect(reverse('user_detail', args=[user.id]))

    return render(request, 'index/user_detail.html', {'user': user})

def update_score(request):
    choice_id = request.POST.get('choice_id')
    score = request.POST.get('score')

    try:
        choice = Choices.objects.get(id=choice_id)
        choice.scores = int(score)
        choice.save()
        return JsonResponse({'status': 'success'})
    except Choices.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Choice not found'})

def change_desc(request):
    if request.method == 'POST':
        new_desc = request.POST.get('desc')  
        user_desc, created = UserDesc.objects.get_or_create(user=request.user)
        user_desc.desc = new_desc
        user_desc.save()

        return redirect('edit_profile')  

    context = {}
    return render(request, 'index/edit-profile.html', context)


def change_date_of_birth(request):
    try:
        user_date_of_birth = request.user.dateofbirth
    except DateOfBirth.DoesNotExist:
        user_date_of_birth = DateOfBirth(user=request.user, date_of_birth=None)
        user_date_of_birth.save()

    try:
        if request.method == 'POST':
            new_date_of_birth = request.POST.get('date_of_birth', None)

            if request.user.dateofbirth.id != user_date_of_birth.id:
                return HttpResponseForbidden("You do not have permission to perform this action.")

            user_date_of_birth.date_of_birth = new_date_of_birth
            user_date_of_birth.save()
            return redirect('edit_profile')

    except Exception as e:
        print(f"An error occurred: {e}")
        return redirect('404') 

    context = {'user': request.user}
    return render(request, 'index/edit-profile.html', context)

def delete_date_of_birth(request):
    user = request.user
    try:
        user_date_of_birth = DateOfBirth.objects.get(user=user)
        user_date_of_birth.delete()
        return redirect('edit_profile')
    except DateOfBirth.DoesNotExist:
        return HttpResponse("Date of Birth not found", status=404)

@login_required
def change_gender(request):
    # Проверяем, имеет ли пользователь профиль с гендером
    try:
        user_gender = request.user.usergender
    except UserGender.DoesNotExist:

        user_gender = UserGender(user=request.user, gender='O')
        user_gender.save()

    if request.method == 'POST':

        new_gender = request.POST.get('gender', 'O')

        if request.user.usergender.id != user_gender.id:
            return HttpResponseForbidden("You do not have permission to perform this action.")

        user_gender.gender = new_gender
        user_gender.save()
        return redirect('edit_profile')

    context = {'user': request.user}
    return render(request, 'index/edit-profile.html', context)

@login_required
def change_username(request):
    if request.method == 'POST':
        new_username = request.POST.get('username')
        if new_username:
            request.user.username = new_username
            request.user.save()
            messages.success(request, 'Username успешно изменен.')
        else:
            messages.error(request, 'Пожалуйста, введите новый username.')

    return redirect('edit_profile')

@login_required
def change_email(request):
    if request.method == 'POST':
        new_email = request.POST.get('email')
        if new_email:
            request.user.email = new_email
            request.user.save()
            messages.success(request, 'Email успешно изменен.')
        else:
            messages.error(request, 'Пожалуйста, введите новый email.')

    return redirect('edit_profile')

def change_profile_image(request):
    if request.method == 'POST':
        image_file = request.FILES.get('imageFile')
        if image_file:
            # Удаляем старое изображение, если оно существует
            old_image = Image.objects.filter(user=request.user).first()
            if old_image:
                old_image.delete()

            # Сохраняем новое изображение
            image = Image(user=request.user)
            image.image.save(image_file.name, ContentFile(image_file.read()))
            image.save()

    return redirect('edit_profile')

def delete_profile_image(request):
    if request.method == 'POST':
        image = Image.objects.filter(user=request.user).first()
        if image:
            image.delete()
    return redirect('edit_profile')


def delete_forms(request):
    if request.method == 'DELETE':
        form_ids = request.POST.getlist('forms[]')
        return JsonResponse({'message': 'Forms deleted successfully'}, status=200)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=400)

def index(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
        
    # Для админов и staff показываем текущий интерфейс
    if request.user.is_superuser or request.user.is_staff:
        forms = Form.objects.all()
        return render(request, "index/index.html", {
            "forms": forms
        })
    
    # Для обычных пользователей показываем активные формы
    active_forms = Form.objects.filter(is_active=True).order_by('-createdAt')
    
    return render(request, "index/home.html", {
        "active_forms": active_forms
    })

def validate_password(password, confirmation):
    if password != confirmation:
        return "Пароли должны совпадать."
    elif len(password) < 8:
        return "Пароль должен содержать как минимум 8 символов."
    elif not any(char.isdigit() for char in password) or not any(char.isalpha() for char in password):
        return "Пароль должен содержать буквы и цифры."
    return None

def login_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect(reverse('index'))
    if request.method == "POST":
        username = request.POST["username"].lower()
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return HttpResponseRedirect(reverse('index'))
        else:
            return render(request, "index/loginregister.html", {
                "message": "Неверное имя пользователя или пароль.",
                "city_choices": UserCity.CITY_CHOICES  # Добавляем city_choices
            })
    return render(request, "index/loginregister.html", {
        "city_choices": UserCity.CITY_CHOICES  # Добавляем city_choices
    })

def register(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect(reverse('index'))
    
    if request.method == "POST":
        username = request.POST["username"].lower()
        password = request.POST["password"]
        email = request.POST["email"]
        confirmation = request.POST["confirmation"]
        region = request.POST.get("region")
        med_center = request.POST.get("med_center")

        if not re.match(r'^[a-zA-Z0-9]+$', username):
            return render(request, "index/loginregister.html", {
                "message": "Имя пользователя может содержать только латинские буквы и цифры.",
                "city_choices": UserCity.CITY_CHOICES
            })

        error_message = validate_password(password, confirmation)
        if error_message:
            return render(request, "index/loginregister.html", {
                "message": error_message,
                "city_choices": UserCity.CITY_CHOICES
            })

        if User.objects.filter(email=email).exists():
            return render(request, "index/loginregister.html", {
                "message": "Email уже занят.",
                "city_choices": UserCity.CITY_CHOICES
            })

        try:
            user = User.objects.create_user(username=username, password=password, email=email)
            
            # Создаем запись о городе
            user_city = UserCity.objects.create(user=user, city=region)
            
            # Создаем запись о медцентре
            UserMed.objects.create(user=user, med_center=med_center)
            
            login(request, user)
            return HttpResponseRedirect(reverse('index'))
        except IntegrityError:
            return render(request, "index/loginregister.html", {
                "message": "Имя пользователя уже занято",
                "city_choices": UserCity.CITY_CHOICES
            })
    
    # Для GET-запроса передаем список городов
    return render(request, "index/loginregister.html", {
        'city_choices': UserCity.CITY_CHOICES,
    })

def get_med_centers(request):
    region = request.GET.get('region')
    med_centers = RegionMedCenter.objects.filter(region=region).values('med_center', 'address')
    return JsonResponse(list(med_centers), safe=False)

def logout_view(request):
    logout(request)
    return HttpResponseRedirect(reverse('index'))

def create_form(request):

    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))

    if request.method == "POST":
        data = json.loads(request.body)
        title = data["title"]
        code = ''.join(random.choice(string.ascii_letters + string.digits) for x in range(30))
        choices = Choices(choice = "Вариант 1")
        choices.save()
        question = Questions(question_type = "multiple choice", question= "Без названия", required= False)
        question.save()
        question.choices.add(choices)
        question.save()
        form = Form(code = code, title = title, creator=request.user)
        form.save()
        form.questions.add(question)
        form.save()
        return JsonResponse({"message": "Success", "code": code})
    else:
        return render(request, 'error/404.html')

def edit_form(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    if not request.user.is_superuser and not request.user.is_staff:
        return HttpResponseRedirect(reverse("403"))
    formInfo = Form.objects.filter(code = code)

    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse("404"))
    else: formInfo = formInfo[0]
    return render(request, "index/form.html", {
        "code": code,
        "form": formInfo
    })

def edit_title(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)

    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse("404"))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        if len(data["title"]) > 0:
            formInfo.title = data["title"]
            formInfo.save()
        else:
            formInfo.title = formInfo.title[0]
            formInfo.save()
        return JsonResponse({"message": "Success", "title": formInfo.title})

def edit_description(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)

    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse("404"))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        formInfo.description = data["description"]
        formInfo.save()
        return JsonResponse({"message": "Success", "description": formInfo.description})

def edit_bg_color(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)

    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse("404"))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        formInfo.background_color = data["bgColor"]
        formInfo.save()
        return JsonResponse({"message": "Success", "bgColor": formInfo.background_color})

def edit_text_color(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)

    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse("404"))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        formInfo.text_color = data["textColor"]
        formInfo.save()
        return JsonResponse({"message": "Success", "textColor": formInfo.text_color})

def edit_setting(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code=code)

    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse("404"))
    else:
        formInfo = formInfo[0]
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        formInfo.collect_email = data.get("collect_email", formInfo.collect_email)
        formInfo.is_quiz = data.get("is_quiz", formInfo.is_quiz)
        formInfo.authenticated_responder = data.get("authenticated_responder", formInfo.authenticated_responder)
        formInfo.confirmation_message = data.get("confirmation_message", formInfo.confirmation_message)
        formInfo.edit_after_submit = data.get("edit_after_submit", formInfo.edit_after_submit)
        formInfo.allow_view_score = data.get("allow_view_score", formInfo.allow_view_score)
        formInfo.limit_ip = data.get("limit_ip", formInfo.limit_ip)
        formInfo.submit_limit = data.get("submit_limit", formInfo.submit_limit)
        formInfo.is_single_form = data.get("is_single_form", formInfo.is_single_form)
        formInfo.is_active = data.get("is_active", formInfo.is_active)
        formInfo.save()
        return JsonResponse({'message': "Success"})

def delete_form(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)

    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse("404"))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "DELETE":

        for i in formInfo.questions.all():
            for j in i.choices.all():
                j.delete()
            i.delete()
        for i in Responses.objects.filter(response_to = formInfo):
            for j in i.response.all():
                j.delete()
            i.delete()
        formInfo.delete()
        return JsonResponse({'message': "Success"})

def edit_question(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    
    form_info = Form.objects.filter(code=code).first()

    if not form_info:
        return HttpResponseRedirect(reverse('404'))
    
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            
            question_id = data.get("id")
            question = Questions.objects.filter(id=question_id).first()
            
            if not question:
                return HttpResponseRedirect(reverse("404"))
            
            question.question = data.get("question", question.question)
            question.question_type = data.get("question_type", question.question_type)
            question.required = data.get("required", question.required)
            question.is_list = data.get("is_list", question.is_list)
            question.is_skip = data.get("is_skip", question.is_skip)
            question.is_negative = data.get("is_negative", question.is_negative)
            if "score" in data:
                question.score = data["score"]
            if "answer_key" in data:
                question.answer_key = data["answer_key"]
            
            question.save()
            
            return JsonResponse({'message': "Success"})
        
        except json.JSONDecodeError as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def edit_choice(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        choice_id = data["id"]
        choice = Choices.objects.filter(id = choice_id)
        if choice.count() == 0:
            return HttpResponseRedirect(reverse("404"))
        else: choice = choice[0]
        choice.choice = data["choice"]
        if(data.get('is_answer')): choice.is_answer = data["is_answer"]
        choice.save()
        return JsonResponse({'message': "Success"})


def add_choice(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))

    form_info = Form.objects.filter(code=code)

    if form_info.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else:
        form_info = form_info[0]

    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))

    if request.method == "POST":
        data = json.loads(request.body)
        choice_number = form_info.questions.get(pk=data["question"]).choices.count() + 1
        choice_text = f"Вариант {choice_number}"

        choice = Choices(choice=choice_text)
        choice.save()

        form_info.questions.get(pk=data["question"]).choices.add(choice)
        form_info.save()

        return JsonResponse({"message": "Success", "choice": choice.choice, "id": choice.id})

def remove_choice(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        choice = Choices.objects.filter(pk = data["id"])
        if choice.count() == 0:
            return HttpResponseRedirect(reverse("404"))
        else: choice = choice[0]
        choice.delete()
        return JsonResponse({"message": "Success"})

def get_choice(request, code, question):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "GET":
        question = Questions.objects.filter(id = question)
        if question.count() == 0: return HttpResponseRedirect(reverse('404'))
        else: question = question[0]
        choices = question.choices.all()
        choices = [{"choice":i.choice, "is_answer":i.is_answer, "id": i.id} for i in choices]
        return JsonResponse({"choices": choices, "question": question.question, "question_type": question.question_type, "question_id": question.id})

from django.db.models import Max

def add_question(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        # Получаем максимальный существующий order
        max_order = formInfo.questions.aggregate(Max('order'))['order__max']
        new_order = (max_order or -1) + 1  # Если вопросов нет, начинаем с 0
        
        choices = Choices(choice = "Вариант 1")
        choices.save()
        question = Questions(
            question_type = "multiple choice", 
            question = "Без названия", 
            required = False,
            order = new_order  # Устанавливаем новый порядок
        )
        question.save()
        question.choices.add(choices)
        question.save()
        formInfo.questions.add(question)
        formInfo.save()
        return JsonResponse({'question': {
            'question': "Без названия", 
            "question_type": "multiple choice", 
            "required": False, 
            "id": question.id
        },
        "choices": {"choice": "Вариант 1", "is_answer": False, 'id': choices.id}})

def delete_question(request, code, question):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))

    # Получаем форму по коду
    formInfo = Form.objects.filter(code=code).first()
    if not formInfo:
        return HttpResponseRedirect(reverse('404'))

    # Проверка прав пользователя
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))

    if request.method == "DELETE":
        # Получаем вопрос по ID
        question = Questions.objects.filter(id=question).first()
        if not question:
            return HttpResponseRedirect(reverse("404"))

        # Удаляем связанные варианты выбора
        question.choices.all().delete()

        # Удаляем сам вопрос
        question.delete()

        return JsonResponse({"message": "Success"})

def score(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if not formInfo.is_quiz:
        return HttpResponseRedirect(reverse("edit_form", args = [code]))
    else:
        return render(request, "index/score.html", {
            "form": formInfo
        })

def edit_score(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if not formInfo.is_quiz:
        return HttpResponseRedirect(reverse("edit_form", args = [code]))
    else:
        if request.method == "POST":
            data = json.loads(request.body)
            question_id = data["question_id"]
            question = formInfo.questions.filter(id = question_id)
            if question.count() == 0:
                return HttpResponseRedirect(reverse("edit_form", args = [code]))
            else: question = question[0]
            score = data["score"]
            if score == "": score = 0
            question.score = score
            question.save()
            return JsonResponse({"message": "Success"})

def answer_key(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if not formInfo.is_quiz:
        return HttpResponseRedirect(reverse("edit_form", args = [code]))
    else:
        if request.method == "POST":
            data = json.loads(request.body)
            question = Questions.objects.filter(id = data["question_id"])
            if question.count() == 0: return HttpResponseRedirect(reverse("edit_form", args = [code]))
            else: question = question[0]
            if question.question_type == "short" or question.question_type == "paragraph" or question.question_type == "range slider":
                question.answer_key = data["answer_key"]
                question.save()
            else:
                for i in question.choices.all():
                    i.is_answer = False
                    i.save()
                if question.question_type == "multiple choice":
                    choice = question.choices.get(pk = data["answer_key"])
                    choice.is_answer = True
                    choice.save()
                else:
                    for i in data["answer_key"]:
                        choice = question.choices.get(id = i)
                        choice.is_answer = True
                        choice.save()
                question.save()
            return JsonResponse({'message': "Success"})

def feedback(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)

    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if not formInfo.is_quiz:
        return HttpResponseRedirect(reverse("edit_form", args = [code]))
    else:
        if request.method == "POST":
            data = json.loads(request.body)
            question = formInfo.questions.get(id = data["question_id"])
            question.feedback = data["feedback"]
            question.save()
            return JsonResponse({'message': "Success"})

def view_form(request, code):
    formInfo = Form.objects.filter(code=code).first()
    if not formInfo:
        return HttpResponseRedirect(reverse("404"))
        
    med_centers = RegionMedCenter.objects.all().order_by('region', 'med_center')
    
    return render(request, "index/view_form.html", {
        "form": formInfo,
        "med_centers": med_centers
    })

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

from datetime import datetime, date, time
from django.utils import timezone
import pytz

def submit_form(request, code):
    formInfo = Form.objects.filter(code=code).first()
    
    if not formInfo:
        return HttpResponseRedirect(reverse('404'))
    
    if formInfo.authenticated_responder and not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    
    if request.method == "POST":
        client_ip = get_client_ip(request)
        
        # Проверяем ограничения IP и лимиты отправки
        if formInfo.limit_ip:
            existing_response = Responses.objects.filter(response_to=formInfo, responder_ip=client_ip).order_by('-createdAt').first()
            if existing_response and datetime.now() - existing_response.createdAt < timedelta(hours=24):
                return render(request, "index/form_response.html", {
                    "form": formInfo,
                    "code": code,
                    "message": "Вы уже отправили ответ на эту форму в последние 24 часа."
                })

        if formInfo.submit_limit:
            existing_response = Responses.objects.filter(response_to=formInfo, responder=request.user).order_by('-createdAt').first()
            if existing_response and datetime.now() - existing_response.createdAt < timedelta(hours=24):
                return render(request, "index/form_response.html", {
                    "form": formInfo,
                    "code": code,
                    "message": "Вы уже отправили ответ на эту форму в последние 24 часа."
                })
        
        code = ''.join(random.choice(string.ascii_letters + string.digits) for x in range(20))
        
        if request.user.is_authenticated and request.user == formInfo.creator:
            # Получаем пользовательские данные
            custom_email = request.POST.get("custom_email")
            custom_gender = request.POST.get("custom_gender")
            custom_city = request.POST.get("custom_city")
            custom_med = request.POST.get("custom_med")
            custom_birth_date = request.POST.get("custom_birth_date")
            custom_username = request.POST.get("custom_username")
            custom_submit_date = request.POST.get("custom_submit_date")
            custom_submit_time = request.POST.get("custom_submit_time")

            # Создаем базовый объект response БЕЗ сохранения
            response = Responses(
                response_code=code,
                response_to=formInfo,
                responder_ip=client_ip,
                responder=request.user
            )

            # Устанавливаем пользовательские данные
            response.responder_email = custom_email if custom_email else request.user.email
            response.responder_gender = custom_gender if custom_gender else (request.user.gender_info.gender if hasattr(request.user, 'gender_info') else None)
            response.responder_city = custom_city if custom_city else (request.user.city_info.city if hasattr(request.user, 'city_info') else None)
            response.responder_med = custom_med if custom_med else (request.user.med_info.med_center if hasattr(request.user, 'med_info') else None)
            response.responder_birth_date = custom_birth_date if custom_birth_date else (request.user.date_info.date_of_birth if hasattr(request.user, 'date_info') else None)
            response.responder_username = custom_username if custom_username else request.user.username

            # Устанавливаем пользовательскую дату и время отправки
            if custom_submit_date and custom_submit_time:
                try:
                    submit_date = datetime.strptime(custom_submit_date, '%Y-%m-%d').date()
                    submit_time = datetime.strptime(custom_submit_time, '%H:%M').time()
                    custom_datetime = datetime.combine(submit_date, submit_time)
                    tz = pytz.timezone('Asia/Yekaterinburg')
                    response.createdAt = tz.localize(custom_datetime)
                except (ValueError, TypeError):
                    response.createdAt = timezone.now()
            else:
                response.createdAt = timezone.now()

            # Вычисляем возраст на основе даты рождения
            if response.responder_birth_date:
                today = datetime.now().date()
                birthdate = datetime.strptime(str(response.responder_birth_date), '%Y-%m-%d').date()
                response.responder_age = today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))

            # Теперь сохраняем объект response
            response.save()

        elif request.user.is_authenticated and formInfo.authenticated_responder:
            response = Responses(
                response_code=code,
                response_to=formInfo,
                responder_ip=client_ip,
                responder=request.user
            )
            response.save()
        elif formInfo.collect_email:
            response = Responses(
                response_code=code,
                response_to=formInfo,
                responder_ip=client_ip,
                responder_email=request.POST.get("email-address")
            )
            response.save()

        # Сохраняем ответы на вопросы
        for i in request.POST:
            if i in ["csrfmiddlewaretoken", "email-address", "custom_email", "custom_gender", 
                    "custom_city", "custom_med", "custom_birth_date", "custom_username",
                    "custom_submit_date", "custom_submit_time"] or i.startswith("is_skipped_"):
                continue
            
            try:
                question_id = int(i)
            except ValueError:
                continue
            
            question = formInfo.questions.get(id=question_id)
            for j in request.POST.getlist(i):
                is_skipped = request.POST.get(f'is_skipped_{question.id}') == 'True'
                answer = Answer(answer=j, answer_to=question, is_skipped=is_skipped)
                answer.save()
                response.response.add(answer)

        response.save()
        
        return render(request, "index/form_response.html", {
            "form": formInfo,
            "code": code
        })

def calculate_average_scores(all_responses, formInfo):
    average_scores = {}
    questions = formInfo.questions.prefetch_related('choices').filter(question_type='range slider')

    # Получаем все уникальные медцентры из ответов
    med_centers = RegionMedCenter.objects.all()
    for med_center in med_centers:
        average_scores[med_center.med_center] = {
            "scores": {question.id: 0 for question in questions},
            "total_score": 0,
            "count": 0,
            "max_values": {question.id: question.max_value for question in questions}
        }

    for response in all_responses:
        med_center = response.responder_med
        if med_center in average_scores:
            average_scores[med_center]["count"] += 1
            for question in questions:
                answer = Answer.objects.filter(response=response, answer_to=question.id, is_skipped=False).first()
                if answer:
                    try:
                        average_scores[med_center]["scores"][question.id] += round(float(answer.answer))
                    except ValueError:
                        continue

    for med_center, data in average_scores.items():
        for question_id, total_score in data["scores"].items():
            if data["count"] > 0:
                average_scores[med_center]["scores"][question_id] = round(total_score / data["count"])
        if data["count"] > 0:
            total_possible_score = sum(data["max_values"].values())
            total_achieved_score = sum(data["scores"].values())
            average_scores[med_center]["total_score"] = round((total_achieved_score / total_possible_score) * 100)  # В процентном соотношении

    return average_scores
    
from dateutil.relativedelta import relativedelta
from django.db.models import Avg
from datetime import datetime, timedelta

def responses(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))

    formInfo = get_object_or_404(Form, code=code)

    if not request.user.is_superuser and not request.user.is_staff:
        return HttpResponseRedirect(reverse("403"))

    # Получаем параметры фильтрации
    selected_city = request.GET.get('cities')
    selected_gender = request.GET.get('gender')
    age_min = request.GET.get('age_min')
    age_max = request.GET.get('age_max')
    selected_med_region = request.GET.get('med_region')
    selected_med_center = request.GET.get('med_center')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Базовый запрос
    all_responses = Responses.objects.filter(response_to=formInfo)
    
    # Применяем фильтры
    if selected_city:
        all_responses = all_responses.filter(responder__city_info__city=selected_city)
    
    if selected_gender:
        all_responses = all_responses.filter(responder__gender_info__gender=selected_gender)
    
    if age_min:
        min_birth_year = datetime.now().year - int(age_min)
        all_responses = all_responses.filter(
            responder__date_info__date_of_birth__year__lte=min_birth_year
        )
    
    if age_max:
        max_birth_year = datetime.now().year - int(age_max)
        all_responses = all_responses.filter(
            responder__date_info__date_of_birth__year__gte=max_birth_year
        )

    # Фильтрация по региону медцентра
    if selected_med_region:
        med_centers = RegionMedCenter.objects.filter(region=selected_med_region).values_list('med_center', flat=True)
        all_responses = all_responses.filter(responder_med__in=med_centers)

    # Фильтрация по конкретному медцентру
    if selected_med_center:
        all_responses = all_responses.filter(responder_med=selected_med_center)

    # Добавляем фильтрацию по датам
    if date_from:
        all_responses = all_responses.filter(createdAt__gte=date_from)
    if date_to:
        # Добавляем 1 день к date_to, чтобы включить весь последний день
        date_to = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
        all_responses = all_responses.filter(createdAt__lt=date_to)

    # Получаем выбранные варианты ответов для вопросов
    choiceAnswered = {}
    for question in formInfo.questions.all():
        if question.question_type in ["multiple choice", "checkbox"]:
            selected_choices = request.GET.getlist(f'question-{question.id}')
            if selected_choices:
                choiceAnswered[question.id] = selected_choices
                answers = Answer.objects.filter(
                    response__in=all_responses,
                    answer_to=question,
                    answer__in=selected_choices
                )
                response_ids = answers.values_list('response_id', flat=True)
                all_responses = all_responses.filter(id__in=response_ids)
    
    # Сохраняем параметры фильтрации в сессии для использования при экспорте
    request.session['filter_params'] = {
        'cities': selected_city,
        'gender': selected_gender,
        'age_min': age_min,
        'age_max': age_max,
        'med_region': selected_med_region,
        'med_center': selected_med_center,
        'date_from': date_from,
        'date_to': request.GET.get('date_to'),  # Сохраняем оригинальную дату
        'choices': choiceAnswered
    }
    
    responsesSummary, choiceAnswered, choices_dict = process_questions_and_answers(formInfo)
    
    range_slider_data = get_range_slider_data(formInfo, all_responses)
    
    response_answers = get_response_answers(all_responses, formInfo)
    
    filteredResponsesSummary = get_filtered_response_summary(choiceAnswered, all_responses, formInfo)
    
    user_city_dict = get_user_city_dict(all_responses)
    
    average_data = get_average_data(formInfo, all_responses)
    
    average_scores = calculate_average_scores(all_responses, formInfo)
    average_scores_sorted = sorted(average_scores.items(), key=lambda x: x[1]['total_score'], reverse=True)
    
    med_center_stats = get_med_center_stats(formInfo, all_responses)

    final_scores = calculate_final_scores(request, code)

    return render(request, "index/responses.html", {
        "form": formInfo,
        "responses": all_responses,
        "responsesSummary": responsesSummary,
        "filteredResponsesSummary": filteredResponsesSummary,
        "choices_dict": choices_dict,
        "average_scores_sorted": average_scores_sorted,
        "response_answers": response_answers,
        "city_choices": UserCity.CITY_CHOICES,
        "user_city_dict": user_city_dict,
        "range_slider_data": range_slider_data,
        "average_data": average_data,
        'med_centers': RegionMedCenter.objects.all().values_list('med_center', flat=True).distinct(),
        "average_scores": average_scores,
        "med_center_stats": med_center_stats,
        "final_scores": final_scores,
        'active_forms': Form.objects.filter(is_active=True),
    })

def get_filtered_responses(formInfo, age_min, age_max, selected_gender, selected_cities):
    all_responses = Responses.objects.filter(response_to=formInfo)
    
    # Используем сохраненные статические данные вместо связей
    if age_min:
        all_responses = all_responses.filter(responder_age__gte=int(age_min))
    if age_max:
        all_responses = all_responses.filter(responder_age__lte=int(age_max))
    if selected_gender:
        all_responses = all_responses.filter(responder_gender=selected_gender)
    if selected_cities and selected_cities != ['']:
        all_responses = all_responses.filter(responder_city__in=selected_cities)
    
    return all_responses

def process_questions_and_answers(formInfo):
    responsesSummary = []
    choiceAnswered = {}
    choices_dict = {}

    for question in formInfo.questions.exclude(question_type="title"):
        answers = Answer.objects.filter(answer_to=question.id, is_skipped=False)

        if question.question_type in ["multiple choice", "checkbox"]:
            choiceAnswered[question.id] = choiceAnswered.get(question.id, {})
            for answer in answers:
                try:
                    choice = answer.answer_to.choices.get(id=answer.answer)
                    choices_dict[answer.answer] = choice.choice
                    unique_choice_key = f"{choice.choice}"
                    choiceAnswered[question.id][unique_choice_key] = choiceAnswered[question.id].get(unique_choice_key, 0) + 1
                except Choices.DoesNotExist:
                    continue

        responsesSummary.append({"question": question, "answers": answers})
    
    return responsesSummary, choiceAnswered, choices_dict

def get_range_slider_data(formInfo, all_responses):
    range_slider_data = {}
    end_date = datetime.now()
    start_date = end_date - timedelta(days=365)
    current_month = end_date.replace(day=1)
    months = [(current_month - relativedelta(months=i)).strftime('%B %Y') for i in range(12)]
    months.reverse()

    questions = formInfo.questions.filter(question_type="range slider").exclude(question_type="title")
    question_ids = questions.values_list('id', flat=True)
    responses_for_questions = Answer.objects.filter(
        answer_to__in=question_ids,
        response__in=all_responses,
        response__createdAt__gte=start_date,
        is_skipped=False
    )

    for question in questions:
        responses_for_question = responses_for_questions.filter(answer_to=question.id)
        centers = responses_for_question.values_list('response__responder_med', flat=True).distinct()
        monthly_averages_by_center = {center: [0] * 12 for center in centers}

        for i in range(12):
            month_start = current_month - relativedelta(months=i)
            month_end = month_start + relativedelta(months=1)

            monthly_responses = responses_for_question.filter(
                response__createdAt__gte=month_start,
                response__createdAt__lt=month_end
            ).values('response__responder_med').annotate(avg_value=Avg('answer'))

            for response in monthly_responses:
                center = response['response__responder_med']
                avg_value = response['avg_value']
                monthly_averages_by_center[center][11 - i] = round(avg_value)

        range_slider_data[question.id] = {
            'months': months,
            'averages_by_center': monthly_averages_by_center,
            'max_value': question.max_value
        }

    return range_slider_data

def get_response_answers(all_responses, formInfo):
    response_answers = {}
    for response in all_responses:
        response_answers[response.id] = {}
        for question in formInfo.questions.exclude(question_type="title"):
            if question.question_type == "checkbox":
                answers = Answer.objects.filter(response=response, answer_to=question.id, is_skipped=False)
                selected_choices = []
                for answer in answers:
                    try:
                        choice = answer.answer_to.choices.get(id=answer.answer)
                        selected_choices.append(choice.choice)
                    except Choices.DoesNotExist:
                        continue
                response_answers[response.id][question.id] = selected_choices if selected_choices else "N/A"
            else:
                answer = Answer.objects.filter(response=response, answer_to=question.id, is_skipped=False).first()
                if answer:
                    if question.question_type == "multiple choice":
                        try:
                            choice = answer.answer_to.choices.get(id=answer.answer)
                            response_answers[response.id][question.id] = choice.choice
                        except Choices.DoesNotExist:
                            response_answers[response.id][question.id] = "N/A"
                    else:
                        response_answers[response.id][question.id] = answer.answer
                else:
                    response_answers[response.id][question.id] = "N/A"
    return response_answers

def get_filtered_response_summary(choiceAnswered, all_responses, formInfo):
    # Создаем словарь для хранения результатов
    filteredResponsesSummary = {}
    
    # Инициализируем счетчики для каждого вопроса и варианта ответа
    for question in formInfo.questions.exclude(question_type="title"):
        if question.question_type in ["multiple choice", "checkbox"]:
            filteredResponsesSummary[question.id] = {}
            for choice in question.choices.all():
                filteredResponsesSummary[question.id][choice.choice] = 0

    # Подсчитываем ответы для отфильтрованных респондентов
    for response in all_responses:
        for question in formInfo.questions.exclude(question_type="title"):
            if question.question_type in ["multiple choice", "checkbox"]:
                answers = Answer.objects.filter(
                    response=response, 
                    answer_to=question.id, 
                    is_skipped=False
                )
                
                if question.question_type == "checkbox":
                    for answer in answers:
                        try:
                            choice = question.choices.get(id=answer.answer)
                            filteredResponsesSummary[question.id][choice.choice] += 1
                        except Choices.DoesNotExist:
                            continue
                else:  # multiple choice
                    answer = answers.first()
                    if answer:
                        try:
                            choice = question.choices.get(id=answer.answer)
                            filteredResponsesSummary[question.id][choice.choice] += 1
                        except Choices.DoesNotExist:
                            continue

    return filteredResponsesSummary

def get_user_city_dict(all_responses):
    # Больше не нужно делать дополнительные запросы к UserCity
    user_city_dict = {}
    for response in all_responses:
        user_city_dict[response.id] = response.responder_city
    return user_city_dict

def get_average_data(formInfo, all_responses):
    average_data = {}
    questions = formInfo.questions.prefetch_related('choices').exclude(question_type="title")

    for question in questions:
        if question.question_type == "range slider":
            question_average_data = {}
            # Используем сохраненный med_center вместо связи
            responders = all_responses.values_list('responder_med', flat=True).distinct()

            for responder_med in responders:
                selected_responses = all_responses.filter(responder_med=responder_med)
                range_answers = Answer.objects.filter(
                    response__in=selected_responses,
                    answer_to=question.id,
                    is_skipped=False
                )

                if range_answers.exists():
                    average_value = range_answers.aggregate(Avg('answer'))['answer__avg']
                    normalized_avg = (average_value / question.max_value) * question.max_value if question.max_value != 0 else 0
                    question_average_data[responder_med] = normalized_avg
                else:
                    question_average_data[responder_med] = 0

            average_data[question.id] = question_average_data
    
    return average_data

def get_med_center_stats(formInfo, all_responses):
    med_center_stats = {}
    med_centers = RegionMedCenter.objects.all()

    # Инициализируем статистику используя ID медцентра
    for med_center in med_centers:
        med_center_stats[med_center.med_center] = {
            'total_responses': 0,
            'questions': {},
        }
        for question in formInfo.questions.filter(is_negative=True):
            med_center_stats[med_center.med_center]['questions'][question.id] = 0

    # Подсчитываем статистику
    for response in all_responses:
        med_center = response.responder_med
        if med_center in med_center_stats:
            med_center_stats[med_center]['total_responses'] += 1
            
            # Используем response вместо answers
            for answer in response.response.all():
                if answer.answer_to.is_negative and answer.answer == 'Да':
                    med_center_stats[med_center]['questions'][answer.answer_to.id] += 1

    return med_center_stats

def delete_users(request):
    if request.method == 'POST':
        selected_users = request.POST.getlist('selected_users[]')
        if selected_users:
            User.objects.filter(id__in=selected_users).delete()
            messages.success(request, 'Пользователи успешно удалены.')
        return redirect('user_list')
from index.templatetags.calculate_score import calculate_score
from index.templatetags.calculate_score import calculate_total_score

def export_responses_to_excel(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))

    formInfo = get_object_or_404(Form, code=code)

    if not request.user.is_superuser and not request.user.is_staff:
        return HttpResponseRedirect(reverse("403"))

    # Получаем состояния чекбоксов
    export_as_percentage = request.GET.get('export_as_percentage') == 'true'
    export_total_as_percentage = request.GET.get('export_total_as_percentage') == 'true'
    hidden_columns = request.GET.get('hidden_columns', '').split(',')

    # Получаем параметры фильтрации из сессии
    filter_params = request.session.get('filter_params', {})
    
    # Применяем фильтры
    all_responses = Responses.objects.filter(response_to=formInfo)
    
    # Фильтрация по городу пользователя
    if filter_params.get('cities'):
        all_responses = all_responses.filter(responder_city=filter_params['cities'])
    
    # Фильтрация по полу
    if filter_params.get('gender'):
        all_responses = all_responses.filter(responder_gender=filter_params['gender'])
    
    # Фильтрация по возрасту
    if filter_params.get('age_min'):
        all_responses = all_responses.filter(responder_age__gte=int(filter_params['age_min']))
    
    if filter_params.get('age_max'):
        all_responses = all_responses.filter(responder_age__lte=int(filter_params['age_max']))
    
    # Фильтрация по медцентру
    if filter_params.get('med_center'):
        all_responses = all_responses.filter(responder_med=filter_params['med_center'])

    # Применяем фильтры по выбранным вариантам ответов
    for question_id, selected_choices in filter_params.get('choices', {}).items():
        answers = Answer.objects.filter(
            response__in=all_responses,
            answer_to_id=question_id,
            answer__in=selected_choices
        )
        response_ids = answers.values_list('response_id', flat=True)
        all_responses = all_responses.filter(id__in=response_ids)

    response_answers = {}
    for response in all_responses:
        response_answers[response.id] = {}
        for question in formInfo.questions.exclude(question_type="title"):
            if question.question_type == "checkbox":
                answers = Answer.objects.filter(response=response, answer_to=question.id, is_skipped=False)
                selected_choices = []
                for answer in answers:
                    try:
                        choice = answer.answer_to.choices.get(id=answer.answer)
                        selected_choices.append(choice.choice)
                    except Choices.DoesNotExist:
                        continue
                response_answers[response.id][question.id] = selected_choices if selected_choices else "N/A"
            else:
                answer = Answer.objects.filter(response=response, answer_to=question.id, is_skipped=False).first()
                if answer:
                    if question.question_type == "multiple choice":
                        try:
                            choice = answer.answer_to.choices.get(id=answer.answer)
                            response_answers[response.id][question.id] = choice.choice
                        except Choices.DoesNotExist:
                            response_answers[response.id][question.id] = "N/A"
                    elif question.question_type == "range slider":
                        try:
                            raw_value = float(answer.answer)
                        except ValueError:
                            raw_value = 0.0
                        max_range_value = question.max_value
                        if export_as_percentage:
                            percentage = (raw_value / max_range_value) * 100
                            response_answers[response.id][question.id] = f"{percentage:.2f}%"
                        else:
                            response_answers[response.id][question.id] = raw_value
                    elif question.question_type in ["short", "paragraph"]:
                        response_answers[response.id][question.id] = answer.answer if answer.answer.strip() else "N/A"
                    else:
                        response_answers[response.id][question.id] = answer.answer
                else:
                    response_answers[response.id][question.id] = "N/A"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Responses"

    headers = ["User", "Age", "Gender", "City", "Medical Center", "Submission Date"]
    question_headers = {question.question: question.id for question in formInfo.questions.exclude(question_type="title")}

    headers = [header for header in headers if header.lower() not in hidden_columns]
    for question, question_id in question_headers.items():
        if f"question-{question_id}" not in hidden_columns:
            headers.append(question)
    if "total-score" not in hidden_columns and formInfo.is_quiz:
        headers.append("Total Score (Percentage)" if export_total_as_percentage else "Total Score")

    # Стили для заголовков
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Записываем заголовки и применяем стили
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_num, response in enumerate(all_responses, 2):
        row = []
        responder = response.responder
        row_data = {
            "User": response.responder_username if response.responder else "Anonymous",
            "Age": response.responder_age if response.responder_age else "N/A",
            "Gender": response.responder_gender if response.responder_gender else "N/A",
            "City": response.responder_city if response.responder_city else "N/A",
            "Medical Center": response.responder_med if response.responder_med else "N/A",
            "Submission Date": response.createdAt.strftime("%d.%m.%Y %H:%M")
        }

        for question, question_id in question_headers.items():
            if f"question-{question_id}" not in hidden_columns:
                answer = response_answers[response.id].get(question_id, "N/A")
                row_data[question] = answer

        if "total-score" not in hidden_columns and formInfo.is_quiz:
            total_score = calculate_score(response, formInfo)
            if export_total_as_percentage:
                total_percentage = (total_score / calculate_total_score(formInfo)) * 100
                row_data["Total Score (Percentage)"] = f"{total_percentage:.2f}%"
            else:
                row_data["Total Score"] = total_score

        row = [row_data[col] for col in headers]
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Автоматическая настройка ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="responses.xlsx"'
    return response

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

def export_combined_excel(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))

    formInfo = get_object_or_404(Form, code=code)

    if not request.user.is_superuser and not request.user.is_staff:
        return HttpResponseRedirect(reverse("403"))

    visible_columns = request.GET.get('visible_columns', '').split(',')
    visible_med_centers = request.GET.get('visible_med_centers', '').split(',')
    export_as_percentage = request.GET.get('export_as_percentage') == 'true'
    export_total_as_percentage = request.GET.get('export_total_as_percentage') == 'true'

    # Получаем параметры фильтрации из сессии
    filter_params = request.session.get('filter_params', {})
    
    # Применяем те же фильтры, что и в представлении responses
    all_responses = Responses.objects.filter(response_to=formInfo)
    
    if filter_params.get('cities'):
        all_responses = all_responses.filter(responder__city_info__city=filter_params['cities'])
    
    if filter_params.get('gender'):
        all_responses = all_responses.filter(responder__gender_info__gender=filter_params['gender'])
    
    if filter_params.get('age_min'):
        all_responses = all_responses.filter(responder_age__gte=int(filter_params['age_min']))
    
    if filter_params.get('age_max'):
        all_responses = all_responses.filter(responder_age__lte=int(filter_params['age_max']))

    if filter_params.get('med_region'):
        med_centers = RegionMedCenter.objects.filter(
            region=filter_params['med_region']
        ).values_list('med_center', flat=True)
        all_responses = all_responses.filter(responder_med__in=med_centers)

    if filter_params.get('med_center'):
        all_responses = all_responses.filter(responder_med=filter_params['med_center'])

    # Применяем фильтры по выбранным вариантам ответов
    for question_id, selected_choices in filter_params.get('choices', {}).items():
        answers = Answer.objects.filter(
            response__in=all_responses,
            answer_to_id=question_id,
            answer__in=selected_choices
        )
        response_ids = answers.values_list('response_id', flat=True)
        all_responses = all_responses.filter(id__in=response_ids)

    average_scores = calculate_average_scores(all_responses, formInfo)
    average_scores_sorted = sorted(average_scores.items(), key=lambda x: x[1]['total_score'], reverse=True)
    med_center_stats = get_med_center_stats(formInfo, all_responses)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined Data"

    # Стили для заголовков
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Заголовки для средних оценок
    headers = ["Медицинский центр"]
    for question in formInfo.questions.filter(question_type="range slider"):
        if f'{question.id}' in visible_columns:
            headers.append(f"{question.question}")
    
    # Добавляем заголовки для количества ответов
    for question in formInfo.questions.filter(question_type__in=["short", "paragraph"]):
        if f'answ-{question.id}' in visible_columns:
            headers.append(f"Кол-во ответов: {question.question}")
    
    if 'main-value' in visible_columns:
        headers.append("Общая оценка")
    if 'total-responses' in visible_columns:
        headers.append("Общее количество ответов")

    # Записываем заголовки и применяем стили
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)

    # Заполняем данные
    row_num = 2
    for med_center, scores in average_scores_sorted:
        if med_center not in visible_med_centers:
            continue
        
        row = [med_center]
        
        # Добавляем значения range slider
        for question in formInfo.questions.filter(question_type="range slider"):
            if f'{question.id}' in visible_columns:
                raw_value = scores['scores'].get(question.id, "N/A")
                if raw_value != "N/A" and export_as_percentage:
                    max_value = question.max_value
                    percentage = (raw_value / max_value) * 100
                    row.append(f"{percentage:.2f}%")
                else:
                    row.append(raw_value)
        
        # Добавляем количество ответов для каждого вопроса
        for question in formInfo.questions.filter(question_type__in=["short", "paragraph"]):
            if f'answ-{question.id}' in visible_columns:
                answers_count = med_center_stats[med_center]['questions'].get(question.id, 0)
                row.append(answers_count)
        
        # Добавляем общую оценку
        if 'main-value' in visible_columns:
            total_score = scores['total_score']
            if export_total_as_percentage:
                total_score = f"{total_score:.2f}%"
            row.append(total_score)
        
        if 'total-responses' in visible_columns:
            row.append(med_center_stats[med_center]['total_responses'])
        
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        row_num += 1

    # Настраиваем ширину столбцов
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Combined_Data_{formInfo.code}.xlsx'
    wb.save(response)

    return response

def retrieve_checkbox_choices(response, question):
    checkbox_answers = []

    answers = Answer.objects.filter(answer_to=question, response=response, is_skipped=False)
    for answer in answers:
        selected_choice_ids = answer.answer.split(',')  
        selected_choices = Choices.objects.filter(pk__in=selected_choice_ids)
        checkbox_answers.append([choice.choice for choice in selected_choices])

    return checkbox_answers

def exportcsv(request,code):
    formInfo = Form.objects.filter(code = code)
    formInfo = formInfo[0]
    responses=Responses.objects.filter(response_to = formInfo)
    questions = formInfo.questions.all()


    http_response = HttpResponse()
    http_response['Content-Type'] = 'text/csv'
    http_response['Content-Disposition'] = f'attachment; filename={formInfo.title}.csv'
    writer = csv.writer(http_response)
    header = ['Response Code', 'Responder', 'Responder Email','Responder_ip']

    for question in questions:
        header.append(question.question)

    writer.writerow(header)

    for response in responses:
        response_data = [
        response.response_code,
        response.responder_username if response.responder else 'Anonymous',
        response.responder_email if response.responder_email else '',
        response.responder_ip if response.responder_ip else ''
    ]
        for question in questions:
            answer = Answer.objects.filter(answer_to=question, response=response, is_skipped=False).first()


            if  question.question_type not in ['multiple choice','checkbox']:
                response_data.append(answer.answer if answer else '')
            elif question.question_type == "multiple choice":
                response_data.append(answer.answer_to.choices.get(id = answer.answer).choice if answer else '')
            elif question.question_type == "checkbox":
                if answer and question.question_type == 'checkbox':
                    checkbox_choices = retrieve_checkbox_choices(response,answer.answer_to)
                    response_data.append(checkbox_choices)

        print(response_data)
        writer.writerow(response_data)

    return http_response

def response(request, code, response_code):
    formInfo = Form.objects.filter(code=code)
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else:
        formInfo = formInfo[0]
    if not formInfo.allow_view_score:
        if not request.user.is_staff:
            return HttpResponseRedirect(reverse("403"))

    total_score = 0
    score = 0
    responseInfo = Responses.objects.filter(response_code=response_code)
    if responseInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else:
        responseInfo = responseInfo[0]

    if formInfo.is_quiz:
        for question in formInfo.questions.all():
            if question.question_type == "multiple choice":
                max_choice_score = max([choice.scores for choice in question.choices.all()])
                total_score += max_choice_score + question.score
            elif question.question_type == "checkbox":
                choices_total_score = sum([choice.scores for choice in question.choices.all()])
                total_score += choices_total_score + question.score
            else:
                total_score += question.score

        _temp = []
        for response in responseInfo.response.all():
            if response.answer_to.question_type in ["short", "paragraph"]:
                if response.answer == response.answer_to.answer_key:
                    score += response.answer_to.score
            elif response.answer_to.question_type == "multiple choice":
                answerKey = None
                choice_score = 0
                for choice in response.answer_to.choices.all():
                    if choice.is_answer:
                        answerKey = choice.id
                    if choice.id == int(response.answer):
                        choice_score = choice.scores
                if answerKey is not None and int(answerKey) == int(response.answer):
                    score += response.answer_to.score + choice_score
            elif response.answer_to.question_type == "checkbox" and response.answer_to.pk not in _temp:
                answers = []
                answer_keys = []
                choice_scores_sum = 0
                selected_scores_sum = 0
                for resp in responseInfo.response.filter(answer_to__pk=response.answer_to.pk):
                    answers.append(int(resp.answer))
                    for choice in resp.answer_to.choices.all():
                        if choice.is_answer and choice.pk not in answer_keys:
                            answer_keys.append(choice.pk)
                        if choice.pk == int(resp.answer):
                            selected_scores_sum += choice.scores
                    _temp.append(response.answer_to.pk)
                if set(answers) == set(answer_keys):
                    score += response.answer_to.score
                score += selected_scores_sum

    return render(request, "index/response.html", {
        "form": formInfo,
        "response": responseInfo,
        "score": score,
        "total_score": total_score
    })

def export_final_scores(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))

    formInfo = get_object_or_404(Form, code=code)
    if not request.user.is_superuser and not request.user.is_staff:
        return HttpResponseRedirect(reverse("403"))

    # Получаем активные формы и данные
    active_forms = Form.objects.filter(is_active=True)
    final_scores = calculate_final_scores(request, code)

    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Итоговые оценки"

    # Создаем заголовки
    headers = ["Медицинский центр"]
    for form in active_forms:
        headers.append(form.title)
    headers.extend(["Количество жалоб", "Процент влияния жалоб", "Итоговая оценка"])

    # Применяем стили к заголовкам
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        # Устанавливаем ширину столбца
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Заполняем данные
    row = 2
    for med_center, data in final_scores.items():
        col = 1
        # Медцентр
        ws.cell(row=row, column=col, value=med_center)
        col += 1
        
        # Оценки по формам
        for form in active_forms:
            score = data['forms'].get(form.title, 'N/A')
            if score != 'N/A':
                ws.cell(row=row, column=col, value=float(score))
            else:
                ws.cell(row=row, column=col, value=score)
            col += 1
        
        # Количество жалоб
        ws.cell(row=row, column=col, value=data['negative_count'])
        col += 1
        
        # Процент влияния жалоб (делим на 100, так как значение уже в процентах)
        percentage_cell = ws.cell(row=row, column=col, value=data['negative_percentage'] / 100)
        percentage_cell.number_format = '0%'
        col += 1
        
        # Итоговая оценка
        ws.cell(row=row, column=col, value=data['total_score'])
        
        row += 1

    # Создаем HTTP ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Final_Scores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    # Сохраняем файл
    wb.save(response)
    return response

def edit_response(request, code, response_code):
    formInfo = Form.objects.filter(code=code).first()
    if not formInfo:
        return HttpResponseRedirect(reverse('404'))
    
    response = Responses.objects.filter(response_code=response_code, response_to=formInfo).first()
    if not response:
        return HttpResponseRedirect(reverse('404'))
    
    if formInfo.authenticated_responder:
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse("login"))
        if response.responder != request.user:
            return HttpResponseRedirect(reverse('403'))
    
    if request.method == "POST":
        if formInfo.authenticated_responder and not response.responder:
            response.responder = request.user
            response.save()
        if formInfo.collect_email:
            response.responder_email = request.POST["email-address"]
            response.save()
        for i in response.response.all():
            i.delete()
        for i in request.POST:
            if i == "csrfmiddlewaretoken" or i == "email-address":
                continue
            question = formInfo.questions.get(id=i)
            for j in request.POST.getlist(i):
                is_skipped = request.POST.get(f'is_skipped_{question.id}') == 'True'
                answer = Answer(answer=j, answer_to=question, is_skipped=is_skipped)
                answer.save()
                response.response.add(answer)
                response.save()
        if formInfo.is_quiz:
            return HttpResponseRedirect(reverse("response", args=[formInfo.code, response.response_code]))
        else:
            return render(request, "index/form_response.html", {
                "form": formInfo,
                "code": response.response_code
            })
    return render(request, "index/edit_response.html", {
        "form": formInfo,
        "response": response
    })

def contact_form_template(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        code = ''.join(random.choice(string.ascii_letters + string.digits) for x in range(30))
        name1 = Choices(choice="Вариант 1")
        name1.save()
        name = Questions(question_type = "short", question= "ФИО", required= True)
        name.save()
        name.choices.add(name1)
        name.save()
        email1 = Choices(choice="Вариант 2")
        email1.save()
        email = Questions(question_type="short", question = "Электронная почта", required = True)
        email.save()
        email.choices.add(email1)
        email.save()
        address1 = Choices(choice="Вариант 2")
        address1.save()
        address = Questions(question_type="paragraph", question="Адрес", required = True)
        address.save()
        address.choices.add(address1)
        address.save()
        phone1 = Choices(choice="Вариант 2")
        phone1.save()
        phone = Questions(question_type="short", question="Номер телефона", required = False)
        phone.save()
        phone.choices.add(phone1)
        phone.save()
        comments1 = Choices(choice="Вариант 2")
        comments1.save()
        comments = Questions(question_type = "paragraph", question = "Комментарии", required = False)
        comments.save()
        comments.choices.add(comments1)
        comments.save()
        form = Form(code = code, title = "Контактная информация", creator=request.user, background_color="#e2eee0", allow_view_score = False, edit_after_submit = True)
        form.save()
        form.questions.add(name)
        form.questions.add(email)
        form.questions.add(address)
        form.questions.add(phone)
        form.questions.add(comments)
        form.save()
        return JsonResponse({"message": "Успешно", "code": code})

def customer_feedback_template(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        code = ''.join(random.choice(string.ascii_letters + string.digits) for x in range(30))
        comment = Choices(choice = "Комментарии")
        comment.save()
        question = Choices(choice = "Вопросы")
        question.save()
        bug = Choices(choice = "Баг-репорт")
        bug.save()
        feature = Choices(choice = "Запрос функции")
        feature.save()
        feedback_type = Questions(question = "Тип отзыва", question_type="multiple choice", required=False)
        feedback_type.save()
        feedback_type.choices.add(comment)
        feedback_type.choices.add(bug)
        feedback_type.choices.add(question)
        feedback_type.choices.add(feature)
        feedback_type.save()
        feedback1 = Choices(choice="Вариант 1")
        feedback1.save()
        feedback = Questions(question = "Отзыв", question_type="paragraph", required=True)
        feedback.save()
        feedback.choices.add(feedback1)
        feedback.save()
        suggestion1 = Choices(choice="Вариант 1")
        suggestion1.save()
        suggestion = Questions(question = "Предложения для улучш��ния", question_type="paragraph", required=False)
        suggestion.save()
        suggestion.choices.add(suggestion1)
        suggestion.save()
        name1 = Choices(choice="Вариант 1")
        name1.save()
        name = Questions(question = "ФИО", question_type="short", required=False)
        name.save()
        name.choices.add(name1)
        name.save()
        email1 = Choices(choice="Вариант 2")
        email1.save()
        email = Questions(question= "E-mail", question_type="short", required=False)
        email.save()
        email.choices.add(email1)
        email.save()
        form = Form(code = code, title = "Форма обратной связи", creator=request.user, background_color="#e2eee0", confirmation_message="Thanks so much for giving us feedback!",
        description = "Мы будем рады услышать ваши мысли или отзывы о том, как мы можем улучшить ваш опыт!", allow_view_score = False, edit_after_submit = True)
        form.save()
        form.questions.add(feedback_type)
        form.questions.add(feedback)
        form.questions.add(suggestion)
        form.questions.add(name)
        form.questions.add(email)
        return JsonResponse({"message": "Успешно", "code": code})

def event_registration_template(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        code = ''.join(random.choice(string.ascii_letters + string.digits) for x in range(30))
        name1 = Choices(choice="Вариант 1")
        name1.save()
        name = Questions(question="ФИО", question_type= "short", required=True)
        name.save()
        name.choices.add(name1)
        name.save()
        email1 = Choices(choice="Вариант 2")
        email1.save()
        email = Questions(question = "E-mail", question_type="short", required=True)
        email.save()
        email.choices.add(email1)
        email.save()
        organization1 = Choices(choice="Вариант 3")
        organization1.save()
        organization = Questions(question = "Организация", question_type= "short", required=True)
        organization.save()
        organization.choices.add(organization1)
        organization.save()
        day1 = Choices(choice="Дня 1")
        day1.save()
        day2 = Choices(choice= "Дня 2")
        day2.save()
        day3 = Choices(choice= "Дня 3")
        day3.save()
        day = Questions(question="В какие дни вы будете присутствовать?", question_type="checkbox", required=True)
        day.save()
        day.choices.add(day1)
        day.choices.add(day2)
        day.choices.add(day3)
        day.save()
        dietary_none = Choices(choice="Нет")
        dietary_none.save()
        dietary_vegetarian = Choices(choice="Вегетарианское")
        dietary_vegetarian.save()
        dietary_kosher = Choices(choice="Кошерные")
        dietary_kosher.save()
        dietary_gluten = Choices(choice = "Без глютена")
        dietary_gluten.save()
        dietary = Questions(question = "Ограничения в питании", question_type="multiple choice", required = True)
        dietary.save()
        dietary.choices.add(dietary_none)
        dietary.choices.add(dietary_vegetarian)
        dietary.choices.add(dietary_gluten)
        dietary.choices.add(dietary_kosher)
        dietary.save()
        accept_agreement = Choices(choice = "Да")
        accept_agreement.save()
        agreement = Questions(question = "Я понимаю, что мне придется заплатить [ДЕНЬГИ] по прибытии", question_type="checkbox", required=True)
        agreement.save()
        agreement.choices.add(accept_agreement)
        agreement.save()
        form = Form(code = code, title = "Регистрация на мероприятие", creator=request.user, background_color="#fdefc3",
        confirmation_message="Мы приняли вашу регистрацию.\n\
Укажите здесь другую информацию.\n\
\n\
Сохраните ссылку ниже, по которой можно редактировать свою регистрацию вплоть до даты ее закрытия.",
        description = "Дата проведения: 4-6 января 2023 г.\n\
Адрес мероприятия: 123 'Ваша улица', 'Ваш город', Проспект 12345\n\
Свяжитесь с нами по телефону (123)-456-7890 и/или example@example.com", edit_after_submit=True, allow_view_score=False)
        form.save()
        form.questions.add(name)
        form.questions.add(email)
        form.questions.add(organization)
        form.questions.add(day)
        form.questions.add(dietary)
        form.questions.add(agreement)
        form.save()
        return JsonResponse({"message": "Успешно", "code": code})

def social_survey_template(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    if not request.user.is_superuser:
        return HttpResponseRedirect(reverse("403"))

    if request.method == "POST":
        code = ''.join(random.choice(string.ascii_letters + string.digits) for x in range(30))

        def create_question(question_text, question_type, max_value=None, choices=None, is_list=False, is_negative=False):
            question = Questions(
                question=question_text,
                question_type=question_type,
                required=True,
                max_value=max_value if max_value is not None else None,
                is_list=is_list,
                is_negative=is_negative
            )
            question.save()
            if choices:
                for choice_text in choices:
                    choice = Choices(choice=choice_text)
                    choice.save()
                    question.choices.add(choice)
            question.save()
            return question

        questions = [
            ("Оценка качества работы регистратора", "range slider", 39),
            ("Визит в процедурный кабинет", "range slider", 24),
            ("Визит к врачу", "range slider", 40),
            ("Звонок в колл-центр", "range slider", 31),
            ("Звонок в МЦ", "range slider", 30),
            ("Состояние МЦ", "range slider", 30),
            ("Запрос в чате", "range slider", 20),
            ("Прикреплены ли вы к Медцентру который оцениваете?", "multiple choice", None, ["Да", "Нет"]),
            ("Способ оплаты медицинских услуг", "multiple choice", None, ["Оплачиваю услуги самостоятельно", "Покрываю свои затраты через медицинскую страховку"]),
            ("Оценка удовлетворенности", "range slider", 100),
            ("Жалобы по качеству обслуживания", "paragraph"),
            ("Предложения по улучшению качества обслуживания", "paragraph")
        ]

        created_questions = [create_question(*q) for q in questions]

        form = Form(
            code=code,
            title="Оценка медцентра",
            creator=request.user,
            background_color="#e2eee0",
            allow_view_score=False,
            edit_after_submit=True
        )
        form.save()

        for question in created_questions:
            form.questions.add(question)
        form.save()

        return JsonResponse({"message": "Успешно", "code": code})

def delete_responses(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser and not request.user.is_staff:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "DELETE":
        responses = Responses.objects.filter(response_to = formInfo)
        for response in responses:
            for i in response.response.all():
                i.delete()
            response.delete()
        return JsonResponse({"message": "Успешно"})

def FourZeroThree(request, exception=None):
    return render(request, 'error/403.html', status=403)

def FourZeroFour(request, exception=None):
    return render(request, 'error/404.html', status=404)

def user_list(request):
    users = User.objects.all()
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    return render(request, 'index/user-list.html', {'users': users})

def user_detail(request, pk):
    user = get_object_or_404(User, pk=pk)
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    context = {
        'user_city': UserCity.objects.get_or_create(user=user)[0],
        'user_med': UserMed.objects.get_or_create(user=user)[0],
        'city_choices': UserCity.CITY_CHOICES,
        'med_centers': RegionMedCenter.objects.all().order_by('region', 'med_center'),
        'user': user
    }
    return render(request, 'index/user_detail.html', context)

def contact_us(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    return render(request, 'index/contactUs.html')

def edit_profile(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))

    if request.method == 'POST':
        new_city = request.POST.get('City')
        if new_city is not None:
            user_city, created = UserCity.objects.get_or_create(user=request.user)
            user_city.city = new_city
            user_city.save()
            return redirect('edit_profile')

    context = {
        'user_city': UserCity.objects.get_or_create(user=request.user)[0],
        'user_med': UserMed.objects.get_or_create(user=request.user)[0],
        'city_choices': UserCity.CITY_CHOICES
    }
    return render(request, 'index/edit-profile.html', context)

def view_profile(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    return render(request, 'index/view-profile.html')

def form_list_view(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    if request.method == 'POST':
        selected_forms = request.POST.getlist('selected_forms[]')

        if selected_forms:
            Form.objects.filter(id__in=selected_forms).delete()

    forms = Form.objects.filter()
    if not request.user.is_superuser and not request.user.is_staff:
        return HttpResponseRedirect(reverse("403"))
    return render(request, 'index/form_list.html', {'forms': forms})

@csrf_exempt  # Добавляем декоратор, если его нет
def update_max_value(request, question_id):
    if not request.user.is_authenticated or not request.user.is_superuser:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
        
    if request.method == 'POST':
        try:
            question = Questions.objects.get(pk=question_id)
            data = json.loads(request.body)
            max_value = data.get('max_value')
            
            if max_value is not None:
                try:
                    max_value = int(max_value)  # Убедимся, что значение целое число
                    question.max_value = max_value
                    question.save()
                    return JsonResponse({'message': 'Max value updated successfully', 'max_value': max_value})
                except ValueError:
                    return JsonResponse({'error': 'Invalid max value format'}, status=400)
            else:
                return JsonResponse({'error': 'Max value not provided'}, status=400)
                
        except Questions.DoesNotExist:
            return JsonResponse({'error': 'Question not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)

def update_question_order(request, code):
    if not request.user.is_authenticated or not request.user.is_superuser:
        return JsonResponse({"error": "Unauthorized"}, status=403)
        
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            question_orders = data.get("questionOrders", [])
            
            # Получаем форму
            form = Form.objects.get(code=code)
            
            # Обновляем порядок для каждого вопроса
            for order_data in question_orders:
                question_id = order_data.get("id")
                new_order = order_data.get("order")
                
                # Проверяем, что вопрос принадлежит этой форме
                if form.questions.filter(id=question_id).exists():
                    question = Questions.objects.get(id=question_id)
                    question.order = new_order
                    question.save()
            
            return JsonResponse({"message": "Success"})
        except Form.DoesNotExist:
            return JsonResponse({"error": "Form not found"}, status=404)
        except Exception as e:
            print(f"Error updating question order: {str(e)}")  # Для отладки
            return JsonResponse({"error": str(e)}, status=400)
            
    return JsonResponse({"error": "Method not allowed"}, status=405)

def calculate_final_scores(request, code):
    active_forms = Form.objects.filter(is_active=True)
    med_centers = RegionMedCenter.objects.all()
    final_scores = {
        mc.med_center: {
            'forms': {},
            'total_score': 0,
            'negative_count': 0,
            'negative_percentage': 100
        } for mc in med_centers
    }
    
    for form in active_forms:
        # Используем сохраненные данные медцентра
        all_responses = Responses.objects.filter(response_to=form)
        average_scores = calculate_average_scores(all_responses, form)
        med_center_stats = get_med_center_stats(form, all_responses)
        
        for med_center in med_centers:
            med_center_name = med_center.med_center
            if med_center_name in average_scores:
                form_score = average_scores[med_center_name]['total_score']
                
                # Используем ID медцентра для получения статистики
                negative_count = sum(med_center_stats[med_center_name]['questions'].values())
                final_scores[med_center_name]['negative_count'] += negative_count
                
                if form_score > 0:
                    final_scores[med_center_name]['forms'][form.title] = form_score
    
    # Вычисляем процент влияния жалоб и итоговую оценку
    for med_center in final_scores:
        negative_count = final_scores[med_center]['negative_count']
        
        if negative_count == 0:
            percentage = 100
        elif negative_count == 1:
            percentage = 80
        elif negative_count == 2:
            percentage = 60
        elif negative_count == 3:
            percentage = 40
        elif negative_count == 4:
            percentage = 20
        else:
            percentage = 0
            
        final_scores[med_center]['negative_percentage'] = percentage
        
        form_scores = final_scores[med_center]['forms'].values()
        if form_scores:
            avg_score = sum(form_scores) / len(form_scores)
            final_scores[med_center]['total_score'] = (avg_score * percentage) / 100
        else:
            final_scores[med_center]['total_score'] = 0
    
    final_scores = dict(sorted(final_scores.items(), key=lambda x: x[1]['total_score'], reverse=True))
    return final_scores